import openpyxl
from pprint import pprint

wb = openpyxl.load_workbook('id.xlsx')
sh = wb['Sheet']
dct = {}
for i in range(1, 578):
    id = sh.cell(i,1).value
    model = sh.cell(i,2).value
    brand = sh.cell(i,3).value
    name = model + ', ' + brand
    dct[name] = id

with open('autolist.txt', 'r', encoding='utf-8') as f:
    lst = f.readlines()
res = []
for item in lst:
    name = item.split('|')[1].strip()
    res.append(item.strip() + ' | ' + str(dct[name]))
with open('autolist.txt', 'w', encoding='utf-8') as f:
    for i in res:
        f.write(i + '\n')
