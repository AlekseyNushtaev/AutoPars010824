import csv

import openpyxl
from regions.stavropol import *
from regions.volgograd import *
from regions.surgut import *
from regions.krasnodar import *
from regions.moscow import *
from regions.chelyabinsk import *
from regions.cheboksari import *
from regions.ufa import *
from regions.ekaterinburg import *
from regions.tumen import *
from regions.saratov import *
from regions.samara import *
from regions.yaroslavl import *
from regions.kemerovo import *
from regions.kazan import *
from regions.spb import *
from regions.omsk import *
from regions.novosib import *
from regions.krasnoyarsk import *
from regions.himki import *
from regions.toliati import *
from regions.moscow_dmitrovka import *


async def parser_dmitrovka(dct_up, browser):
    try:
        res_1 = await fili_auto(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://fili-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await plusauto_moscow(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://plusauto.moscow error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'fili-auto.ru_price'
    sheet.cell(row=1, column=7).value = 'fili-auto.ru'
    sheet.cell(row=1, column=8).value = 'plusauto.moscow_price'
    sheet.cell(row=1, column=9).value = 'plusauto.moscow'
    lst_res = [res_1, res_2]
    lst_res_name = [res_1_name, res_2_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/dmitrovka.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/dmitrovka.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_omsk(dct_up, browser):
    # try:
    #     res_1 = await vostoc_ac(dct_up)
    # except Exception as e:
    #     res_1 = []
    #     await bot.send_message(CHANEL_ID, 'https://vostok-ac.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_2 = await center_irtysh(dct_up, browser)
    # except Exception as e:
    #     res_2 = []
    #     await bot.send_message(CHANEL_ID, 'https://center-irtysh.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_3 = await astella_cars(dct_up)
    # except Exception as e:
    #     res_3 = []
    #     await bot.send_message(CHANEL_ID, 'https://astella-cars.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    try:
        res_1 = await irtysh_avtosalon(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://irtysh-avtosalon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_2 = await omsk_cars(dct_up, browser)
    # except Exception as e:
    #     res_2 = []
    #     await bot.send_message(CHANEL_ID, 'https://omsk-cars.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_3 = await omsk_carso(dct_up)
    # except Exception as e:
    #     res_3 = []
    #     await bot.send_message(CHANEL_ID, 'https://omsk.carso.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    # res_2_name = [x[0] for x in res_2]
    # res_3_name = [x[0] for x in res_3]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'irtysh-avtosalon.ru_price'
    sheet.cell(row=1, column=7).value = 'irtysh-avtosalon.ru'
    # sheet.cell(row=1, column=8).value = 'omsk-cars.ru_price'
    # sheet.cell(row=1, column=9).value = 'omsk-cars.ru'
    # sheet.cell(row=1, column=10).value = 'omsk.carso.ru_price'
    # sheet.cell(row=1, column=11).value = 'omsk.carso.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/omsk.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/omsk.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_spb(dct_up, browser):
    try:
        res_1 = await spb_carso(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://spb.carso.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_2 = await credit_cars_spb(dct_up)
    # except Exception as e:
    #     res_2 = []
    #     await bot.send_message(CHANEL_ID, 'https://credit-cars-spb.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_3 = await renault_nnov(dct_up, browser)
    # except Exception as e:
    #     res_3 = []
    #     await bot.send_message(CHANEL_ID, 'https://renault-nnov.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_4 = await altimus_auto(dct_up, browser)
    # except Exception as e:
    #     res_4 = []
    #     await bot.send_message(CHANEL_ID, 'https://altimus-auto.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_5 = await avalon_newspb(dct_up, browser)
    # except Exception as e:
    #     res_5 = []
    #     await bot.send_message(CHANEL_ID, 'https://avalon-newspb.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await ac_neva(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://ac-neva.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'spb.carso.ru_price'
    sheet.cell(row=1, column=7).value = 'spb.carso.ru'
    sheet.cell(row=1, column=8).value = 'ac-neva.ru_price'
    sheet.cell(row=1, column=9).value = 'ac-neva.ru'
    lst_res = [res_1, res_2]
    lst_res_name = [res_1_name, res_2_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/spb.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/spb.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def spb2(dct_up, browser):
    try:
        res_1 = await autosalon_arena_req(dct_up)
    except:
        try:
            res_1 = await autosalon_arena(dct_up, browser)
        except Exception as e:
            res_1 = []
            await bot.send_message(CHANEL_ID, 'https://autosalon-arena.ru error')
            await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'autosalon-arena.ru_price'
    sheet.cell(row=1, column=7).value = 'autosalon-arena.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/spb2.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/spb2.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_ekaterinburg(dct_up, browser):
    try:
        res_1 = await uu_stocks(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://uu-stoks.ru/auto error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await atc_gagarin(dct_up)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://atc-gagarin.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await primeauto_ekb(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://primeauto-ekb.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await swmauto_dealer(dct_up, browser)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://swmauto-dealer.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'uu-stoks.ru_price'
    sheet.cell(row=1, column=7).value = 'uu-stoks.ru'
    sheet.cell(row=1, column=8).value = 'atc-gagarin.ru_price'
    sheet.cell(row=1, column=9).value = 'atc-gagarin.ru'
    sheet.cell(row=1, column=10).value = 'primeauto-ekb.ru_price'
    sheet.cell(row=1, column=11).value = 'primeauto-ekb.ru'
    sheet.cell(row=1, column=12).value = 'swmauto-dealer.ru_price'
    sheet.cell(row=1, column=13).value = 'swmauto-dealer.ru'
    lst_res = [res_1, res_2, res_3, res_4]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/ekaterinburg.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/ekaterinburg.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_saratov(dct_up):
    try:
        res_1 = await saratov_avtohous(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://saratov-avtohous.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await saratov_autospot(dct_up)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://saratov.autospot.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await saratov_asavtomotors(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://saratov.asavtomotors.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await saratov_avtosalon(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://saratov.avtosalon.shop error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await saratov_autosalon(dct_up)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://saratov.autosalon.shop error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'saratov-avtohous.ru_price'
    sheet.cell(row=1, column=7).value = 'saratov-avtohous.ru'
    sheet.cell(row=1, column=8).value = 'saratov.autospot.ru_price'
    sheet.cell(row=1, column=9).value = 'saratov.autospot.ru'
    sheet.cell(row=1, column=10).value = 'saratov.asavtomotors.ru_price'
    sheet.cell(row=1, column=11).value = 'saratov.asavtomotors.ru'
    sheet.cell(row=1, column=12).value = 'saratov.avtosalon.shop_price'
    sheet.cell(row=1, column=13).value = 'saratov.avtosalon.shop'
    sheet.cell(row=1, column=14).value = 'saratov.autosalon.shop_price'
    sheet.cell(row=1, column=15).value = 'saratov.autosalon.shop'
    lst_res = [res_1, res_2, res_3, res_4, res_5]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/saratov.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/saratov.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_kemerovo(dct_up, browser):
    try:
        res_1 = await autonova_nkz(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://autonova-nkz.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await autocenter_kemerevo(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://autocenter-kemerovo.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await center_carplaza(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://center-carplaza.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await carplaza_avtosalon(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://carplaza-avtosalon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await carplaza_ac(dct_up)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://carplaza-ac.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_6 = await kemerevo_autochina(dct_up)
    except Exception as e:
        res_6 = []
        await bot.send_message(CHANEL_ID, 'https://kemerovo-autochina.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_7 = await lada_kemerovo_42(dct_up)
    except Exception as e:
        res_7 = []
        await bot.send_message(CHANEL_ID, 'https://lada-kemerovo42.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5 + res_6 + res_7
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_6_name = [x[0] for x in res_6]
    res_7_name = [x[0] for x in res_7]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'autonova-nkz.ru_price'
    sheet.cell(row=1, column=7).value = 'autonova-nkz.ru'
    sheet.cell(row=1, column=8).value = 'autocenter-kemerovo.ru_price'
    sheet.cell(row=1, column=9).value = 'autocenter-kemerovo.ru'
    sheet.cell(row=1, column=10).value = 'center-carplaza.ru_price'
    sheet.cell(row=1, column=11).value = 'center-carplaza.ru'
    sheet.cell(row=1, column=12).value = 'carplaza-avtosalon.ru_price'
    sheet.cell(row=1, column=13).value = 'carplaza-avtosalon.ru'
    sheet.cell(row=1, column=14).value = 'carplaza-ac.ru_price'
    sheet.cell(row=1, column=15).value = 'carplaza-ac.ru'
    sheet.cell(row=1, column=16).value = 'kemerovo-autochina.ru_price'
    sheet.cell(row=1, column=17).value = 'kemerovo-autochina.ru'
    sheet.cell(row=1, column=18).value = 'lada-kemerovo42.ru_price'
    sheet.cell(row=1, column=19).value = 'lada-kemerovo42.ru'
    lst_res = [res_1, res_2, res_3, res_4, res_5, res_6, res_7]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name, res_6_name, res_7_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/kemerovo.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/kemerovo.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_yaroslavl(dct_up, browser):
    try:
        res_1 = await ac_magistral(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://ac-magistral.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'ac-magistral.ru_price'
    sheet.cell(row=1, column=7).value = 'ac-magistral.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/yaroslavl.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/yaroslavl.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_samara(dct_up, browser):
    try:
        res_1 = await ace_auto_63(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://ace-auto-63.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'ace-auto-63.ru_price'
    sheet.cell(row=1, column=7).value = 'ace-auto-63.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/samara.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/samara.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_volgograd(dct_up, browser):
    try:
        res_1 = await volga_2024_auto(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://volga-2024-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await avto_volga(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://avto-volga-2025.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'volga-2024-auto.ru_price'
    sheet.cell(row=1, column=7).value = 'volga-2024-auto.ru'
    sheet.cell(row=1, column=8).value = 'avto-volga-2025.ru_price'
    sheet.cell(row=1, column=9).value = 'avto-volga-2025.ru'
    lst_res = [res_1, res_2]
    lst_res_name = [res_1_name, res_2_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/volgograd.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/volgograd.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_stavropol(dct_up):
    try:
        res_1 = await autocenter_stav(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://autocenter-stav.ru/auto/ error')
        await bot.send_message(ADMIN_ID, str(e))

    res = res_1
    res_1_name = [x[0] for x in res_1]

    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'autocenter-stav.ru_price'
    sheet.cell(row=1, column=7).value = 'autocenter-stav.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/stavropol.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/stavropol.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_krasnodar(dct_up, browser):
    try:
        res_1 = await avangard_yug(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://avangard-yug.ru/auto error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await rostov_avto(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://rostov-avto-1.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await maximum_auto_credit(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://maximum-auto-credit.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await avtosvoboda_krd(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://avtosvoboda-krd.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]


    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'avangard-yug.ru_price'
    sheet.cell(row=1, column=7).value = 'avangard-yug.ru'
    sheet.cell(row=1, column=8).value = 'rostov-avto-1.ru_price'
    sheet.cell(row=1, column=9).value = 'rostov-avto-1.ru'
    sheet.cell(row=1, column=10).value = 'maximum-auto-credit.ru_price'
    sheet.cell(row=1, column=11).value = 'maximum-auto-credit.ru'
    sheet.cell(row=1, column=12).value = 'avtosvoboda-krd.ru_price'
    sheet.cell(row=1, column=13).value = 'avtosvoboda-krd.ru'

    lst_res = [res_1, res_2, res_3, res_4]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/krasnodar.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/krasnodar.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)



async def parser_cheboksari(dct_up, browser):
    try:
        res_1 = await avto_alyans(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://alyans-auto.ru/auto/auto.html error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'alyans-auto.ru_price'
    sheet.cell(row=1, column=7).value = 'alyans-auto.ru'

    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/cheboksari.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/cheboksari.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_surgut(dct_up, browser):
    try:
        res_1 = await sibir(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://sibir-morots.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await ruauto_s(dct_up)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://ruauto-s.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await fast_autodealer(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://fast-autodealer.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'sibir-morots.ru_price'
    sheet.cell(row=1, column=7).value = 'sibir-morots.ru'
    sheet.cell(row=1, column=8).value = 'ruauto-s.ru_price'
    sheet.cell(row=1, column=9).value = 'ruauto-s.ru'
    sheet.cell(row=1, column=10).value = 'fast-autodealer.ru_price'
    sheet.cell(row=1, column=11).value = 'fast-autodealer.ru'

    lst_res = [res_1, res_2, res_3]
    lst_res_name = [res_1_name, res_2_name, res_3_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/surgut.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/surgut.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_tumen(dct_up, browser):
    # try:
    #     res_1 = await avtosrf(dct_up)
    # except Exception as e:
    #     res_1 = []
    #     await bot.send_message(CHANEL_ID, 'https://avtosrf5-11.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    try:
        res_1 = await bazis_motor(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://bazis-motors.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_3 = await tumen_car(dct_up)
    # except Exception as e:
    #     res_3 = []
    #     await bot.send_message(CHANEL_ID, 'https://tumen-car.ru/ error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_3 = await autocentr_city(dct_up)
    # except Exception as e:
    #     res_3 = []
    #     await bot.send_message(CHANEL_ID, 'https://autocentr-city.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_5 = await autotumen(dct_up)
    # except Exception as e:
    #     res_5 = []
    #     await bot.send_message(CHANEL_ID, 'https://autotumen.ru/ error')
    #     await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await sibtrackt_salon(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://sibtrackt-salon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await tumen_salon(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://tumen-salon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await leks_avto_credit(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://leks-avto-credit.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    # res_5_name = [x[0] for x in res_5]
    # res_6_name = [x[0] for x in res_6]
    # res_7_name = [x[0] for x in res_7]
    # res_8_name = [x[0] for x in res_8]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    # sheet.cell(row=1, column=6).value = 'avtosrf5-11.ru_price'
    # sheet.cell(row=1, column=7).value = 'avtosrf5-11.ru'
    sheet.cell(row=1, column=6).value = 'bazis-motors.ru_price'
    sheet.cell(row=1, column=7).value = 'bazis-motors.ru'
    # sheet.cell(row=1, column=10).value = 'tumen-car.ru_price'
    # sheet.cell(row=1, column=11).value = 'tumen-car.ru'
    # sheet.cell(row=1, column=10).value = 'autocentr-city.ru_price'
    # sheet.cell(row=1, column=11).value = 'autocentr-city.ru'
    # sheet.cell(row=1, column=14).value = 'autotumen.ru_price'
    # sheet.cell(row=1, column=15).value = 'autotumen.ru'
    # sheet.cell(row=1, column=12).value = 'avto-trend72.ru_price'
    # sheet.cell(row=1, column=13).value = 'avto-trend72.ru'
    sheet.cell(row=1, column=8).value = 'sibtrackt-salon.ru_price'
    sheet.cell(row=1, column=9).value = 'sibtrackt-salon.ru'
    sheet.cell(row=1, column=10).value = 'tumen-salon.ru_price'
    sheet.cell(row=1, column=11).value = 'tumen-salon.ru'
    sheet.cell(row=1, column=12).value = 'leks-avto-credit.ru_price'
    sheet.cell(row=1, column=13).value = 'leks-avto-credit.ru'

    lst_res = [res_1, res_2, res_3, res_4]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        if res_name[i - 2].split(', ')[1] == 'Jolion':
            sheet.cell(row=i, column=4).value = 935520
        if res_name[i - 2].split(', ')[1] == 'Jolion New':
            sheet.cell(row=i, column=4).value = 959520
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/tumen.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/tumen.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_chelyabinsk(dct_up, browser):
    # try:
    #     res_1 = await avto_graf(dct_up, browser)
    # except Exception as e:
    #     res_1 = []
    #     await bot.send_message(CHANEL_ID, 'https://avto-graf-newcars.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    try:
        res_1 = await auto_graf74(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://auto-graf74.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await saturn2(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://saturn2.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await kcelitauto(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://kcelitauto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_5 = await kc_klassavto(dct_up)
    # except Exception as e:
    #     res_5 = []
    #     await bot.send_message(CHANEL_ID, 'https://kc-klassavto.ru error')
    #     await bot.send_message(ADMIN_ID, str(e))
    # try:
    #     res_6 = await carsklad_174(dct_up)
    # except Exception as e:
    #     res_6 = []
    #     await bot.send_message(CHANEL_ID, 'https://carsklad-174.ru/auto error')
    #     await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'auto-graf74.ru_price'
    sheet.cell(row=1, column=7).value = 'auto-graf74.ru'
    sheet.cell(row=1, column=8).value = 'saturn2.ru_price'
    sheet.cell(row=1, column=9).value = 'saturn2.ru'
    sheet.cell(row=1, column=10).value = 'kcelitauto.ru_price'
    sheet.cell(row=1, column=11).value = 'kcelitauto.ru'
    # sheet.cell(row=1, column=14).value = 'kc-klassavto.ru_price'
    # sheet.cell(row=1, column=15).value = 'kc-klassavto.ru'
    # sheet.cell(row=1, column=16).value = 'carsklad-174.ru_price'
    # sheet.cell(row=1, column=17).value = 'carsklad-174.ru'

    lst_res = [res_1, res_2, res_3]
    lst_res_name = [res_1_name, res_2_name, res_3_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/chelyabinsk.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/chelyabinsk.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_moscow(dct_up, browser):
    try:
        res_1 = await avanta_avto_credit(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://avanta-avto-credit.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await dc_dbr(dct_up)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://dc-dbr.ru/catalog error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await carsmo(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://carsmo.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await warshauto(dct_up, browser)
    except:
        try:
            res_4 = await warshauto(dct_up, browser)
        except Exception as e:
            res_4 = []
            await bot.send_message(CHANEL_ID, 'https://warshauto.ru/ error')
            await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await kosmos_cars(dct_up)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://kosmos-cars.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_6 = await idol_avto(dct_up)
    except Exception as e:
        res_6 = []
        await bot.send_message(CHANEL_ID, 'https://idol-avto.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_7 = await vita_avto(dct_up)
    except Exception as e:
        res_7 = []
        await bot.send_message(CHANEL_ID, 'https://vita-auto.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_8 = await ca_geely(dct_up, browser)
    except Exception as e:
        res_8 = []
        await bot.send_message(CHANEL_ID, 'https://ca-geely.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_9 = await ac_cronos_msk(dct_up, browser)
    except Exception as e:
        res_9 = []
        await bot.send_message(CHANEL_ID, 'https://ac-kronos-msk.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_10 = await autohous_group(dct_up)
    except Exception as e:
        res_10 = []
        await bot.send_message(CHANEL_ID, 'https://avtohous-group.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = (res_1 + res_2 + res_3 + res_4 + res_5 + res_6 + res_7 + res_8 + res_9 + res_10)
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_6_name = [x[0] for x in res_6]
    res_7_name = [x[0] for x in res_7]
    res_8_name = [x[0] for x in res_8]
    res_9_name = [x[0] for x in res_9]
    res_10_name = [x[0] for x in res_10]

    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'avanta-avto-credit.ru_price'
    sheet.cell(row=1, column=7).value = 'avanta-avto-credit.ru'
    sheet.cell(row=1, column=8).value = 'dc-dbr.ru_price'
    sheet.cell(row=1, column=9).value = 'dc-dbr.ru'
    sheet.cell(row=1, column=10).value = 'carsmo.ru_price'
    sheet.cell(row=1, column=11).value = 'carsmo.ru'
    sheet.cell(row=1, column=12).value = 'warshauto.ru_price'
    sheet.cell(row=1, column=13).value = 'warshauto.ru'
    sheet.cell(row=1, column=14).value = 'kosmos-cars.ru_price'
    sheet.cell(row=1, column=15).value = 'kosmos-cars.ru'
    sheet.cell(row=1, column=16).value = 'idol-avto.ru_price'
    sheet.cell(row=1, column=17).value = 'idol-avto.ru'
    sheet.cell(row=1, column=18).value = 'vita-auto.ru_price'
    sheet.cell(row=1, column=19).value = 'vita-auto.ru'
    sheet.cell(row=1, column=20).value = 'ca-geely.ru_price'
    sheet.cell(row=1, column=21).value = 'ca-geely.ru'
    sheet.cell(row=1, column=22).value = 'ac-kronos-msk.ru_price'
    sheet.cell(row=1, column=23).value = 'ac-kronos-msk.ru'
    sheet.cell(row=1, column=24).value = 'avtohous-group.ru_price'
    sheet.cell(row=1, column=25).value = 'avtohous-group.ru'
    lst_res = [res_1, res_2, res_3, res_4, res_5, res_6, res_7, res_8, res_9, res_10]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name, res_6_name, res_7_name, res_8_name,
                    res_9_name, res_10_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i-2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i-2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i-2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/moscow.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/moscow.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_ufa(dct_up, browser):
    try:
        # await bot.send_message(ADMIN_ID, 'begin https://motors-ag.ru/')
        res_1 = await motors_ag(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://motors-ag.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        # await bot.send_message(ADMIN_ID, 'begin https://ufa.masmotors.ru/')
        res_2 = await ufa_masmotors(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://ufa.masmotors.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        # await bot.send_message(ADMIN_ID, 'begin https://avrora-motors.ru/catalog')
        res_3 = await avrora(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://avrora-motors.ru/catalog error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        # await bot.send_message(ADMIN_ID, 'begin https://ufa.autospot.ru/')
        res_4 = await ufa_autospot(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://ufa.autospot.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        # await bot.send_message(ADMIN_ID, 'begin https://alfa-tank.ru/')
        res_5 = await alpha_tank(dct_up, browser)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://alfa-tank.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'motors-ag.ru_price'
    sheet.cell(row=1, column=7).value = 'motors-ag.ru'
    sheet.cell(row=1, column=8).value = 'ufa.masmotors.ru_price'
    sheet.cell(row=1, column=9).value = 'ufa.masmotors.ru'
    sheet.cell(row=1, column=10).value = 'avrora-motors.ru_price'
    sheet.cell(row=1, column=11).value = 'avrora-motors.ru'
    sheet.cell(row=1, column=12).value = 'ufa.autospot.ru_price'
    sheet.cell(row=1, column=13).value = 'ufa.autospot.ru'
    sheet.cell(row=1, column=14).value = 'alfa-tank.ru_price'
    sheet.cell(row=1, column=15).value = 'alfa-tank.ru'
    lst_res = [res_1, res_2, res_3, res_4, res_5]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/ufa.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/ufa.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_kazan(dct_up, browser):
    try:
        res_1 = await lvl_auto(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://lvl-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await globus_auto(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://globus-auto16.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await kazan_avtosalon(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://kazan.avtosalon.shop error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await kanavto(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://kanavto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await dialog_auto(dct_up)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://dialog-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_6 = await level_motors(dct_up, browser)
    except Exception as e:
        res_6 = []
        await bot.send_message(CHANEL_ID, 'https://level-motors16.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_7 = await kazan_avto(dct_up, browser)
    except Exception as e:
        res_7 = []
        await bot.send_message(CHANEL_ID, 'https://kazan-avtomobili-2025.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_8 = await l_auto(dct_up, browser)
    except Exception as e:
        res_8 = []
        await bot.send_message(CHANEL_ID, 'https://l-auto16.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_9 = await newavto_kazan(dct_up)
    except Exception as e:
        res_9 = []
        await bot.send_message(CHANEL_ID, 'https://newavto-kazan.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5 + res_6 + res_7 + res_8 + res_9
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_6_name = [x[0] for x in res_6]
    res_7_name = [x[0] for x in res_7]
    res_8_name = [x[0] for x in res_8]
    res_9_name = [x[0] for x in res_9]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'lvl-auto.ru_price'
    sheet.cell(row=1, column=7).value = 'lvl-auto.ru'
    sheet.cell(row=1, column=8).value = 'globus-auto16.ru_price'
    sheet.cell(row=1, column=9).value = 'globus-auto16.ru'
    sheet.cell(row=1, column=10).value = 'kazan.avtosalon.shop_price'
    sheet.cell(row=1, column=11).value = 'kazan.avtosalon.shop'
    sheet.cell(row=1, column=12).value = 'kanavto.ru_price'
    sheet.cell(row=1, column=13).value = 'kanavto.ru'
    sheet.cell(row=1, column=14).value = 'dialog-auto.ru_price'
    sheet.cell(row=1, column=15).value = 'dialog-auto.ru'
    sheet.cell(row=1, column=16).value = 'level-motors16.ru_price'
    sheet.cell(row=1, column=17).value = 'level-motors16.ru'
    sheet.cell(row=1, column=18).value = 'kazan-avtomobili-2025.ru_price'
    sheet.cell(row=1, column=19).value = 'kazan-avtomobili-2025.ru'
    sheet.cell(row=1, column=20).value = 'l-auto16.ru_price'
    sheet.cell(row=1, column=21).value = 'l-auto16.ru'
    sheet.cell(row=1, column=22).value = 'newavto-kazan.ru_price'
    sheet.cell(row=1, column=23).value = 'newavto-kazan.ru'

    lst_res = [res_1, res_2, res_3, res_4, res_5, res_6, res_7, res_8, res_9]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name, res_6_name, res_7_name, res_8_name, res_9_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/kazan.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/kazan.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_nsk(dct_up, browser):
    try:
        res_1 = await astella_cars(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://astella-cars.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await kia_novo(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://kia-novo.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await lada_novosib(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://lada-novosib.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await sib_autosalon(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'https://sib-autosalon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await ac_azimut(dct_up, browser)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://ac-azimut.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_6 = await nsk_drive(dct_up, browser)
    except Exception as e:
        res_6 = []
        await bot.send_message(CHANEL_ID, 'https://nsk-drive.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_7 = await sibear_auto(dct_up)
    except Exception as e:
        res_7 = []
        await bot.send_message(CHANEL_ID, 'https://sibear-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_8 = await nsk_avtomir(dct_up)
    except Exception as e:
        res_8 = []
        await bot.send_message(CHANEL_ID, 'https://nsk.avtomir.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_9 = await centorauto_nsk(dct_up)
    except Exception as e:
        res_9 = []
        await bot.send_message(CHANEL_ID, 'https://centorauto-nsk.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_10 = await azimuth_auto(dct_up, browser)
    except Exception as e:
        res_10 = []
        await bot.send_message(CHANEL_ID, 'https://azimuth-auto.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5 + res_6 + res_7 + res_8 + res_9 + res_10
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_6_name = [x[0] for x in res_6]
    res_7_name = [x[0] for x in res_7]
    res_8_name = [x[0] for x in res_8]
    res_9_name = [x[0] for x in res_9]
    res_10_name = [x[0] for x in res_10]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'astella-cars.ru_price'
    sheet.cell(row=1, column=7).value = 'astella-cars.ru'
    sheet.cell(row=1, column=8).value = 'kia-novo.ru_price'
    sheet.cell(row=1, column=9).value = 'kia-novo.ru'
    sheet.cell(row=1, column=10).value = 'lada-novosib.ru_price'
    sheet.cell(row=1, column=11).value = 'lada-novosib.ru'
    sheet.cell(row=1, column=12).value = 'sib-autosalon.ru_price'
    sheet.cell(row=1, column=13).value = 'sib-autosalon.ru'
    sheet.cell(row=1, column=14).value = 'ac-azimut.ru_price'
    sheet.cell(row=1, column=15).value = 'ac-azimut.ru'
    sheet.cell(row=1, column=16).value = 'nsk-drive.ru_price'
    sheet.cell(row=1, column=17).value = 'nsk-drive.ru'
    sheet.cell(row=1, column=18).value = 'sibear-auto.ru_price'
    sheet.cell(row=1, column=19).value = 'sibear-auto.ru'
    sheet.cell(row=1, column=20).value = 'nsk.avtomir.ru_price'
    sheet.cell(row=1, column=21).value = 'nsk.avtomir.ru'
    sheet.cell(row=1, column=22).value = 'centorauto-nsk.ru_price'
    sheet.cell(row=1, column=23).value = 'centorauto-nsk.ru'
    sheet.cell(row=1, column=24).value = 'azimuth-auto.ru_price'
    sheet.cell(row=1, column=25).value = 'azimuth-auto.ru'
    lst_res = [res_1, res_2, res_3, res_4, res_5, res_6, res_7, res_8, res_9, res_10]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name, res_6_name, res_7_name, res_8_name,
                    res_9_name, res_10_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/nsk.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/nsk.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_krsk(dct_up, browser):
    try:
        res_1 = await car_avangard(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://car-avangard.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await lada_kras(dct_up)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://lada-kras.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await krasnoyarsk_carso(dct_up, browser)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://krasnoyarsk.carso.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_4 = await autonew_krr(dct_up)
    except Exception as e:
        res_4 = []
        await bot.send_message(CHANEL_ID, 'http://autonew-krr.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_5 = await neokars(dct_up, browser)
    except Exception as e:
        res_5 = []
        await bot.send_message(CHANEL_ID, 'https://neokars.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3 + res_4 + res_5
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_4_name = [x[0] for x in res_4]
    res_5_name = [x[0] for x in res_5]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'car-avangard.ru_price'
    sheet.cell(row=1, column=7).value = 'car-avangard.ru'
    sheet.cell(row=1, column=8).value = 'lada-kras.ru_price'
    sheet.cell(row=1, column=9).value = 'lada-kras.ru'
    sheet.cell(row=1, column=10).value = 'krasnoyarsk.carso.ru_price'
    sheet.cell(row=1, column=11).value = 'krasnoyarsk.carso.ru'
    sheet.cell(row=1, column=12).value = 'autonew-krr.ru_price'
    sheet.cell(row=1, column=13).value = 'autonew-krr.ru'
    sheet.cell(row=1, column=14).value = 'neokars.ru_price'
    sheet.cell(row=1, column=15).value = 'neokars.ru'
    lst_res = [res_1, res_2, res_3, res_4, res_5]
    lst_res_name = [res_1_name, res_2_name, res_3_name, res_4_name, res_5_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/krsk.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/krsk.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_himki(dct_up, browser):
    try:
        res_1 = await autogansa(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://autogansa.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'autogansa.ru_price'
    sheet.cell(row=1, column=7).value = 'autogansa.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/himki.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/himki.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def parser_toliati(dct_up, browser):
    try:
        res_1 = await carplex_avto63(dct_up, browser)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://carplex-avto63.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_2 = await sool_cars(dct_up, browser)
    except Exception as e:
        res_2 = []
        await bot.send_message(CHANEL_ID, 'https://sool-cars.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    try:
        res_3 = await carplex_autosalon(dct_up)
    except Exception as e:
        res_3 = []
        await bot.send_message(CHANEL_ID, 'https://carplex-avtosalon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1 + res_2 + res_3
    res_1_name = [x[0] for x in res_1]
    res_2_name = [x[0] for x in res_2]
    res_3_name = [x[0] for x in res_3]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'carplex-avto63.ru_price'
    sheet.cell(row=1, column=7).value = 'carplex-avto63.ru'
    sheet.cell(row=1, column=8).value = 'sool-cars.ru_price'
    sheet.cell(row=1, column=9).value = 'sool-cars.ru'
    sheet.cell(row=1, column=10).value = 'carplex-avtosalon.ru_price'
    sheet.cell(row=1, column=11).value = 'carplex-avtosalon.ru'
    lst_res = [res_1, res_2, res_3]
    lst_res_name = [res_1_name, res_2_name, res_3_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/toliati.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/toliati.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def kemerovo2(dct_up):
    try:
        res_1 = await carplaza_avtosalon(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://carplaza-avtosalon.ru/ error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'carplaza-avtosalon.ru_price'
    sheet.cell(row=1, column=7).value = 'carplaza-avtosalon.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/kemerovo2.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/kemerovo2.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def toliati2(dct_up):
    try:
        res_1 = await carplex_autosalon(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://carplex-avtosalon.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'carplex-avtosalon.ru_price'
    sheet.cell(row=1, column=7).value = 'carplex-avtosalon.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/toliati2.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/toliati2.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)


async def surgut2(dct_up):
    try:
        res_1 = await fast_autodealer(dct_up)
    except Exception as e:
        res_1 = []
        await bot.send_message(CHANEL_ID, 'https://fast-autodealer.ru error')
        await bot.send_message(ADMIN_ID, str(e))
    res = res_1
    res_1_name = [x[0] for x in res_1]
    res_name = []
    for item in res:
        if item[0] not in res_name:
            res_name.append(item[0])
    res_name.sort()
    wb = openpyxl.Workbook()
    sheet = wb['Sheet']
    dct_id = {}
    with open('autolist.txt', 'r', encoding='utf-8') as f:
        lst = f.readlines()
    for item in lst:
        try:
            dct_id[item.split('|')[1].strip()] = item.split('|')[2].strip()
        except Exception:
            pass
    sheet.cell(row=1, column=1).value = 'id'
    sheet.cell(row=1, column=2).value = 'brand'
    sheet.cell(row=1, column=3).value = 'model'
    sheet.cell(row=1, column=4).value = 'min_price'
    sheet.cell(row=1, column=5).value = 'min_price_url'
    sheet.cell(row=1, column=6).value = 'fast-autodealer.ru_price'
    sheet.cell(row=1, column=7).value = 'fast-autodealer.ru'
    lst_res = [res_1]
    lst_res_name = [res_1_name]
    for i in range(2, len(res_name) + 2):
        try:
            sheet.cell(row=i, column=1).value = dct_id[res_name[i - 2].strip()]
        except Exception:
            sheet.cell(row=i, column=1).value = 'Новая машина, необходимо назначить id'
        sheet.cell(row=i, column=2).value = res_name[i - 2].split(', ')[0]
        sheet.cell(row=i, column=3).value = res_name[i - 2].split(', ')[1]
        dct = {}
        lst = []
        for y in range(len(lst_res_name)):
            if res_name[i - 2] in lst_res_name[y]:
                index = lst_res_name[y].index(res_name[i - 2])
                sheet.cell(row=i, column=6 + y * 2).value = lst_res[y][index][1]
                sheet.cell(row=i, column=7 + y * 2).value = lst_res[y][index][2]
                dct[str(lst_res[y][index][1])] = lst_res[y][index][2]
                lst.append(int(lst_res[y][index][1]))
        sheet.cell(row=i, column=4).value = min(lst)
        sheet.cell(row=i, column=5).value = dct[str(min(lst))]
    wb.save('xlsx/surgut2.xlsx')
    data = []
    for i in range(1, len(res_name) + 1):
        string = []
        for y in range(2, 5):
            string.append(sheet.cell(row=i, column=y).value)
        data.append(string)
    with open('csv/surgut2.csv', 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f)
        writer.writerows(data)
