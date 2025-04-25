import random
from pprint import pprint

import openpyxl
import requests
import fake_headers
import json
import csv
from bot import bot
from config import ADMIN_ID, CHANEL_ID


async def json_maker():
    for region in ['krasnodar', 'moscow', 'stavropol', 'surgut', 'volgograd', 'samara', 'kazan', 'spb', 'omsk', 'himki',
                   'chelyabinsk', 'cheboksari', 'ufa', 'tumen', 'ekaterinburg', 'saratov', 'kemerovo', 'nsk', 'krsk',
                   'toliati', 'dmitrovka', 'kemerovo2', 'toliati2', 'surgut2']:
        if region in ['krsk', 'dmitrovka', 'moscow', 'kemerovo', 'nsk']:
            dct_region = {}
            res_lst = []
            res = {}
            with open(f'csv/{region}.csv', 'r', encoding='utf-8') as csvfile:
                lst_file = list(csv.reader(csvfile))[1:]
                for row in lst_file:
                    name = row[0] + ', ' + row[1]
                    dct_region[name] = int(row[2])
            res_all = []
            wb = openpyxl.load_workbook('id.xlsx')
            sh = wb['Sheet']
            for i in range(1, 1000):
                model_id = sh.cell(i, 4).value
                if model_id:
                    brand = sh.cell(i, 2).value
                    model = sh.cell(i, 3).value
                    name = brand + ', ' + model
                    price_rrc = sh.cell(i, 5).value
                    price_min = dct_region.get(name, 0)
                    res_all.append([model_id, name, price_rrc, price_min])
            res_all.sort(key=lambda x: x[0])
            for cat in ['BAIC', 'Belgee,', 'Changan', 'Chery', 'Chevrolet', 'DFM', 'Datsun', 'EXEED', 'Gac', 'FAW',
                        'Geely',
                        'Great Wall', 'Haval', 'Hyundai', 'JAC', 'JAECOO', 'Jetour', 'Jetta,', 'KAIYI', 'KIA',
                        'Lada, Granta',
                        'Lada, Largus', 'Lada, Niva', 'Lada, Vesta', 'Lada, Xray', 'Lifan', 'Livan', 'Moskvich',
                        'Nissan', 'OMODA', 'Ravon', 'Renault', 'Skoda', 'Solaris,', 'TANK', 'UAZ', 'Volkswagen',
                        'Zotye']:
                lst_cat = []
                for model in res_all:
                    if cat in model[1]:
                        lst_cat.append(model)
                lst_cat.sort(key=lambda x: x[2])
                if lst_cat[0][3] == 0:
                    baza = 0.6
                else:
                    baza = lst_cat[0][3] / lst_cat[0][2]
                for car in lst_cat:
                    price_min_fix = (int(baza * car[2]) // 100) * 100
                    res_lst.append([int(car[0]), car[1], price_min_fix])
            res_lst.sort(key=lambda x: x[0])
            for item in res_lst:
                if str(item[0]) in res.keys():
                    print(item)
                res[str(item[0])] = {'model': item[1], 'min_price': item[2]}
            json_object = json.dumps(res, indent=4)
            with open(f'json/{region}.json', 'w') as f:
                f.write(json_object)

        elif region in ['tumen', 'kemerovo2']:
            dct_region = {}
            res_lst = []
            res = {}
            with open(f'csv/{region}.csv', 'r', encoding='utf-8') as csvfile:
                lst_file = list(csv.reader(csvfile))[1:]
                for row in lst_file:
                    name = row[0] + ', ' + row[1]
                    dct_region[name] = int(row[2])
            wb = openpyxl.load_workbook('id.xlsx')
            sh = wb['Sheet']
            for i in range(1, 1000):
                model_id = sh.cell(i, 4).value
                if model_id:
                    brand = sh.cell(i, 2).value
                    model = sh.cell(i, 3).value
                    name = brand + ', ' + model
                    price_min = dct_region.get(name, 0)
                    price_rrc = sh.cell(i, 5).value
                    if price_min != 0:
                        res_lst.append([model_id, price_min])
                    else:
                        price_min = int(((0.6 * price_rrc) // 100) * 100)
                        res_lst.append([model_id, price_min])
            res_lst.sort()
            for item in res_lst:
                res[str(item[0])] = {'min_price': item[1]}
            json_object = json.dumps(res, indent=4)
            with open(f'json/{region}.json', 'w') as f:
                f.write(json_object)

        else:
            dct_region = {}
            res_lst = []
            res = {}
            with open(f'csv/{region}.csv', 'r', encoding='utf-8') as csvfile:
                lst_file = list(csv.reader(csvfile))[1:]
                for row in lst_file:
                    name = row[0] + ', ' + row[1]
                    dct_region[name] = int(row[2])
            wb = openpyxl.load_workbook('id.xlsx')
            sh = wb['Sheet']
            for i in range(1, 1000):
                model_id = sh.cell(i, 4).value
                if model_id:
                    brand = sh.cell(i, 2).value
                    model = sh.cell(i, 3).value
                    name = brand + ', ' + model
                    price_min = dct_region.get(name, 0)
                    if price_min != 0:
                        res_lst.append([model_id, price_min])
            res_lst.sort()
            for item in res_lst:
                res[str(item[0])] = {'min_price': item[1]}
            json_object = json.dumps(res, indent=4)
            with open(f'json/{region}.json', 'w') as f:
                f.write(json_object)
