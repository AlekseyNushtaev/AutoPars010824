import openpyxl

wb = openpyxl.load_workbook('moscow.xlsx')
sh = wb['Sheet']
lst = []
for i in range(2, 1000):
    try:
        model_id = int(sh.cell(i, 1).value)
        flag = sh.cell(i, 6).value
        if flag:
            lst.append(model_id)
    except:
        break
print(len(lst))
wb = openpyxl.load_workbook('../models_tech.xlsx')
sh = wb['Sheet']
for i in range(2, 249):
    model_id = sh.cell(i, 5).value
    if model_id in lst:
        sh.cell(i, 11).value = 'Да'
wb.save('../models_tech.xlsx')
