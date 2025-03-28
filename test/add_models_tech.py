import openpyxl

wb = openpyxl.load_workbook('../id.xlsx')
sh = wb['Sheet']
res = []
res_id = []
for i in range(1, 1000):
    brand = sh.cell(i, 2).value
    if brand:
        id_model = sh.cell(i, 1).value
        model = str(sh.cell(i, 3).value)
        name = brand + ', ' + model
        res.append(name)
        res_id.append(id_model)
wb = openpyxl.load_workbook('../models_tech.xlsx')
sh = wb['Sheet']
for i in range(1, 248):
    brand = sh.cell(i, 1).value.lower()
    model = str(sh.cell(i, 2).value).lower()
    name = brand + ', ' + str(model)
    for y in range(len(res)):
        if res[y].lower() == name:
            brand_1 = res[y].split(', ')[0]
            model_1 = res[y].split(', ')[1]
            sh.cell(i, 3).value = brand_1
            sh.cell(i, 4).value = model_1
            sh.cell(i, 5).value = res_id[y]
            break
wb.save('../models_tech.xlsx')
