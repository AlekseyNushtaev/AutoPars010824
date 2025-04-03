import json
from pprint import pprint

import openpyxl

dct_id_gen = {}
wb = openpyxl.load_workbook('../id.xlsx')
sh = wb['Sheet']
for i in range(1, 1000):
    mod_id = sh.cell(i, 4).value
    dct_id_gen[mod_id] = sh.cell(i, 7).value
with open('../models_id.json', 'r', encoding='utf-8') as f:
    data = json.load(f)
print(data)

lst_dct = []
for item in data:
    mod_id = item['id']
    brand = item['brand']
    model = item['model'].replace('New', "").replace('2', "").strip()
    lst_dct.append({'id': mod_id, 'brand': brand, 'generation': dct_id_gen[mod_id], 'model': model})
pprint(lst_dct)
with open('../models_id.json', "w", encoding='utf-8') as f:
    json.dump(lst_dct, f, indent=4, ensure_ascii=False)

