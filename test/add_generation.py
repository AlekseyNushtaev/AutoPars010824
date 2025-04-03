from pprint import pprint

import openpyxl

wb = openpyxl.load_workbook(('../models_tech.xlsx'))
sh = wb['Sheet']
res_1 = []
for i in range(2, 249):
    brand = sh.cell(i, 1).value
    model = sh.cell(i, 2).value
    mod_id_my = sh.cell(i, 5).value
    res_1.append([brand, model, mod_id_my])
wb = openpyxl.load_workbook(('../generations.xlsx'))
sh = wb['Лист1']
dct_id_gen = {}
for i in range(2, 249):
    brand = sh.cell(i, 2).value
    model = sh.cell(i, 3).value
    for y in range(len(res_1)):
        if res_1[y][0] == brand and res_1[y][1] == model:
            dct_id_gen[res_1[y][2]] = sh.cell(i, 4).value
            break
pprint(dct_id_gen)



wb = openpyxl.load_workbook(('../id.xlsx'))
sh = wb['Sheet']
for i in range(1, 1000):
    mod_id = sh.cell(i, 4).value
    if mod_id:
        id_my = sh.cell(i, 1).value
        try:
            sh.cell(i, 7).value = dct_id_gen[id_my]
        except:
            pass
wb.save('../id.xlsx')