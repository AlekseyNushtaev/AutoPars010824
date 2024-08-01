import csv

import openpyxl
import requests
import bs4
import fake_headers
from pprint import pprint

from regions.stavropol import autoshop26
from regions.surgut import autosurgut186, profsouz, aspect, sibir

res_all = autoshop26() + autosurgut186() + profsouz() + aspect() + sibir()
res = []
res_name = []
for item in res_all:
    if item[0] not in res_name:
        res.append(item)
        res_name.append(item[0])
res.sort(key=lambda x: x[0])
pprint(res)


# link = 'https://autoyug26.ru/auto/'
# response = requests.get(link, headers.generate())
# html = response.text
# soup = bs4.BeautifulSoup(html, 'lxml')
# cards = soup.find_all(attrs={"class": "catalog--brands-list--brand--model"})
# res_yug26 = []
# for card in cards:
#     data = card.get("data-model").replace('null', 'None').replace('\\', '')
#     dct = eval(data)
#     link = 'https://autoyug26.ru' + card.find("a").get("href")
#     res_yug26.append([dct["brand"] + ', ' + dct["model"], dct["cost"], link])

# res = [x[0] for x in res_shop26]
# res_2 = [x[0] for x in res_yug26]
# for item in res_2:
#     if item not in res:
#         res.append(item)
# res.sort()
# print(len(res))
# res_yug26_name = [x[0] for x in res_yug26]
# res_shop26_name = [x[0] for x in res_shop26]
#
# wb = openpyxl.Workbook()
# sheet = wb['Sheet']
# sheet.cell(row=1, column=1).value = 'brand'
# sheet.cell(row=1, column=2).value = 'model'
# sheet.cell(row=1, column=3).value = 'min_price'
# sheet.cell(row=1, column=4).value = 'min_price_url'
# sheet.cell(row=1, column=5).value = 'autoshop26.ru'
# sheet.cell(row=1, column=6).value = 'autoshop26.ru_price'
# for i in range(2, len(res) + 2):
#     sheet.cell(row=i, column=1).value = res[i-2].split(', ')[0]
#     sheet.cell(row=i, column=2).value = res[i-2].split(', ')[1]
#     index = res_shop26_name.index(res[i-2])
#     sheet.cell(row=i, column=5).value = res_shop26[index][1]
#     sheet.cell(row=i, column=6).value = res_shop26[index][2]
#     index = res_yug26_name.index(res[i-2])
#     sheet.cell(row=i, column=7).value = res_yug26[index][1]
#     sheet.cell(row=i, column=8).value = res_yug26[index][2]
#     if res_yug26[index][1] > res_shop26[index][1]:
#         sheet.cell(row=i, column=3).value = res_yug26[index][1]
#         sheet.cell(row=i, column=4).value = res_yug26[index][2]
#     else:
#         sheet.cell(row=i, column=3).value = res_shop26[index][1]
#         sheet.cell(row=i, column=4).value = res_shop26[index][2]
# wb.save('stavropol.xlsx')
# data = []
# for i in range(1, len(res) + 1):
#     string = []
#     for y in range(1, 5):
#         string.append(sheet.cell(row=i, column=y).value)
#     data.append(string)
# with open('stavropol.csv', 'w', encoding='utf-8') as f:
#     writer = csv.writer(f)
#     writer.writerows(data)

